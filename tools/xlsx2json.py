#!/usr/bin/env python3
"""
xlsx2json.py - 《通货与铁砧》导表工具
将 data/ 目录下的 xlsx 文件转换为 JSON 和 C# 代码

用法: python tools/xlsx2json.py
      或直接运行 data/xlsx2json.exe

目录约定:
  data/             → xlsx 源文件（按系统分子文件夹，如 data/test/testXlsx.xlsx）
  demo/data/        → 导出的 JSON 文件
  demo/scripts/data/generated/ → 导出的 C# 代码（仅 __enum__.cs + TableRecord.cs + TableData.cs + TableManager.cs）

表结构约定:
  第1行: 字段名称
  第2行: 数据类型（int/float/double/string/bool/枚举名/(list#sep=X),T）
  第3行: 字段备注
  第4行起: 数据行（A列为 ## 的行不导出）
"""

import os
import json
import re
import sys
import keyword
import ctypes

_RED = '\033[91m'
_YELLOW = '\033[93m'
_GREEN = '\033[92m'
_RESET = '\033[0m'

if sys.platform == 'win32':
    kernel32 = ctypes.windll.kernel32
    kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)

def _c(text, color):
    return f"{color}{text}{_RESET}"

CS_KEYWORDS = set(keyword.kwlist) | {
    'abstract', 'as', 'base', 'bool', 'break', 'byte', 'case', 'catch', 'char',
    'checked', 'class', 'const', 'continue', 'decimal', 'default', 'delegate',
    'do', 'double', 'else', 'enum', 'event', 'explicit', 'extern', 'false',
    'finally', 'fixed', 'float', 'for', 'foreach', 'goto', 'if', 'implicit',
    'in', 'int', 'interface', 'internal', 'is', 'lock', 'long', 'namespace',
    'new', 'null', 'object', 'operator', 'out', 'override', 'params', 'private',
    'protected', 'public', 'readonly', 'ref', 'return', 'sbyte', 'sealed',
    'short', 'sizeof', 'stackalloc', 'static', 'string', 'struct', 'switch',
    'this', 'throw', 'true', 'try', 'typeof', 'uint', 'ulong', 'unchecked',
    'unsafe', 'ushort', 'using', 'virtual', 'void', 'volatile', 'while',
    'var', 'dynamic', 'async', 'await', 'partial', 'yield', 'nameof',
    'value', 'get', 'set', 'add', 'remove', 'global', 'alias', 'ascending',
    'descending', 'from', 'group', 'into', 'join', 'let', 'on', 'orderby',
    'select', 'where', 'NotImplementedException',
}

VALID_IDENTIFIER_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')

from openpyxl import load_workbook


def validate_enum_name(name, source):
    errors = []
    if not VALID_IDENTIFIER_RE.match(name):
        errors.append(f"[错误] 枚举名 '{name}' 不是合法的C#标识符 ({source})")
    if name in CS_KEYWORDS:
        errors.append(f"[错误] 枚举名 '{name}' 是C#关键字，请修改 ({source})")
    return errors


def validate_enum_member(enum_name, member_name, member_value, source):
    errors = []
    if not member_name:
        errors.append(f"[错误] 枚举 {enum_name} 的值 {member_value} 缺少枚举成员名(B列) ({source})")
    elif not VALID_IDENTIFIER_RE.match(member_name):
        errors.append(f"[错误] 枚举 {enum_name} 的成员名 '{member_name}' 不是合法的C#标识符 ({source})")
    if member_name and member_name in CS_KEYWORDS:
        errors.append(f"[错误] 枚举 {enum_name} 的成员名 '{member_name}' 是C#关键字，请修改 ({source})")
    return errors


def parse_enum_xlsx(xlsx_path):
    file_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    all_errors = []
    enum_defs = {}

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        all_errors.append(f"[错误] 无法打开文件 {file_name}.xlsx: {e}")
        return enum_defs, all_errors

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 1:
            continue

        current_enum = None
        current_members = []

        for row in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value

            if a_val is None:
                continue

            a_str = str(a_val).strip()

            if not a_str:
                continue

            is_number = False
            try:
                int_val = int(float(a_str))
                is_number = True
            except (ValueError, TypeError):
                pass

            if not is_number:
                if current_enum is not None:
                    enum_defs[current_enum] = current_members
                current_enum = a_str
                current_members = []
                source = f"{file_name}.xlsx / {sheet_name} / 第{row}行"
                all_errors.extend(validate_enum_name(current_enum, source))
            else:
                if current_enum is None:
                    source = f"{file_name}.xlsx / {sheet_name} / 第{row}行"
                    all_errors.append(f"[错误] 第{row}行A列值为数值 '{a_str}'，但之前没有枚举名定义 ({source})")
                    continue

                member_name = str(b_val).strip() if b_val is not None else ''
                comment = str(c_val).strip() if c_val is not None else ''
                source = f"{file_name}.xlsx / {sheet_name} / {current_enum} / 第{row}行"
                all_errors.extend(validate_enum_member(current_enum, member_name, a_str, source))
                current_members.append({
                    'name': member_name,
                    'value': int_val,
                    'comment': comment,
                })

        if current_enum is not None:
            enum_defs[current_enum] = current_members

    wb.close()

    seen_names = {}
    for enum_name, members in enum_defs.items():
        name_set = set()
        value_set = set()
        for m in members:
            if m['name'] in name_set:
                all_errors.append(f"[错误] 枚举 {enum_name} 中存在重复的成员名 '{m['name']}'")
            name_set.add(m['name'])
            if m['value'] in value_set:
                all_errors.append(f"[错误] 枚举 {enum_name} 中存在重复的值 {m['value']} (成员 '{m['name']}')")
            value_set.add(m['value'])
        if enum_name in seen_names:
            all_errors.append(f"[错误] 枚举名 '{enum_name}' 重复定义 (在 {seen_names[enum_name]} 和当前Sheet中)")
        seen_names[enum_name] = enum_name

    return enum_defs, all_errors


def parse_bean_xlsx(xlsx_path):
    file_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    all_errors = []
    bean_defs = {}

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        all_errors.append(f"[错误] 无法打开文件 {file_name}.xlsx: {e}")
        return bean_defs, all_errors

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 2:
            continue

        fields_header_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and str(val).strip() == '*fields':
                fields_header_col = col
                break

        if fields_header_col is None:
            continue

        field_name_col = fields_header_col
        field_type_col = fields_header_col + 2

        current_bean = None
        current_fields = []

        for row in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            e_val = ws.cell(row=row, column=5).value

            is_comment = (a_val is not None and str(a_val).strip() == '##')

            if is_comment:
                if current_bean is not None:
                    bean_defs[current_bean['full_name']] = current_bean
                current_bean = None
                current_fields = []
                continue

            b_str = str(b_val).strip() if b_val is not None else ''

            if b_str and BEAN_FULL_NAME_RE.match(b_str):
                if current_bean is not None:
                    bean_defs[current_bean['full_name']] = current_bean
                sep = str(e_val).strip() if e_val is not None else ','
                current_bean = {
                    'full_name': b_str,
                    'sep': sep,
                    'fields': [],
                }
                current_fields = []
                source = f"{file_name}.xlsx / {sheet_name} / {b_str}"
                short_name = b_str.split('.')[-1]
                if not VALID_IDENTIFIER_RE.match(short_name):
                    all_errors.append(f"[错误] Bean名 '{short_name}' 不是合法的C#标识符 ({source})")
                if short_name in CS_KEYWORDS:
                    all_errors.append(f"[错误] Bean名 '{short_name}' 是C#关键字，请修改 ({source})")
                fname_val = ws.cell(row=row, column=field_name_col).value
                ftype_val = ws.cell(row=row, column=field_type_col).value
                if fname_val is not None:
                    fname = str(fname_val).strip()
                    ftype = str(ftype_val).strip() if ftype_val is not None else 'int'
                    if fname:
                        current_bean['fields'].append({
                            'name': fname,
                            'type': ftype,
                        })
                continue

            if current_bean is not None:
                fname_val = ws.cell(row=row, column=field_name_col).value
                ftype_val = ws.cell(row=row, column=field_type_col).value
                if fname_val is not None:
                    fname = str(fname_val).strip()
                    ftype = str(ftype_val).strip() if ftype_val is not None else 'int'
                    if fname:
                        source = f"{file_name}.xlsx / {sheet_name} / {current_bean['full_name']} / {fname}"
                        if not VALID_IDENTIFIER_RE.match(fname):
                            all_errors.append(f"[错误] Bean {current_bean['full_name']} 的字段名 '{fname}' 不是合法的C#标识符 ({source})")
                        if fname in CS_KEYWORDS:
                            all_errors.append(f"[错误] Bean {current_bean['full_name']} 的字段名 '{fname}' 是C#关键字，请修改 ({source})")
                        current_bean['fields'].append({
                            'name': fname,
                            'type': ftype,
                        })

        if current_bean is not None:
            bean_defs[current_bean['full_name']] = current_bean

    wb.close()

    seen_names = {}
    for full_name, bean_def in bean_defs.items():
        short_name = full_name.split('.')[-1]
        if short_name in seen_names:
            all_errors.append(f"[错误] Bean短名 '{short_name}' 重复定义 ({seen_names[short_name]} 和 {full_name})")
        seen_names[short_name] = full_name

        fname_set = set()
        for f in bean_def['fields']:
            if f['name'] in fname_set:
                all_errors.append(f"[错误] Bean {full_name} 中存在重复的字段名 '{f['name']}'")
            fname_set.add(f['name'])

    return bean_defs, all_errors


def validate_bean_refs(bean_defs, enum_defs):
    all_errors = []
    all_bean_short_names = {fn.split('.')[-1]: fn for fn in bean_defs}

    for full_name, bean_def in bean_defs.items():
        for f in bean_def['fields']:
            ftype = f['type']
            type_info = parse_type(ftype, bean_defs)
            if type_info['kind'] == 'enum':
                if ftype not in enum_defs:
                    all_errors.append(f"[错误] Bean {full_name} 字段 '{f['name']}' 引用了未定义的枚举 '{ftype}'")
            elif type_info['kind'] == 'bean':
                ref_name = ftype
                if ref_name not in bean_defs and ref_name not in all_bean_short_names:
                    all_errors.append(f"[错误] Bean {full_name} 字段 '{f['name']}' 引用了未定义的Bean '{ref_name}'")
            elif type_info['kind'] == 'list':
                elem = type_info['element_type']
                if elem not in BASIC_TYPES and elem not in enum_defs and elem not in bean_defs and elem not in all_bean_short_names:
                    if not VALID_IDENTIFIER_RE.match(elem):
                        all_errors.append(f"[错误] Bean {full_name} 字段 '{f['name']}' 的list元素类型 '{elem}' 不是合法标识符")

    visited = set()
    path_set = set()

    def check_cycle(full_name):
        if full_name in path_set:
            return True
        if full_name in visited:
            return False
        path_set.add(full_name)
        if full_name in bean_defs:
            for f in bean_defs[full_name]['fields']:
                ftype = f['type']
                ti = parse_type(ftype, bean_defs)
                ref = None
                if ti['kind'] == 'bean':
                    ref = ti['type']
                elif ti['kind'] == 'list' and ti['element_type'] not in BASIC_TYPES and ti['element_type'] not in enum_defs:
                    ref = ti['element_type']
                if ref:
                    resolved = all_bean_short_names.get(ref, ref)
                    if resolved in bean_defs and check_cycle(resolved):
                        all_errors.append(f"[错误] Bean {full_name} 字段 '{f['name']}' 存在循环引用")
                        return True
        path_set.remove(full_name)
        visited.add(full_name)
        return False

    for full_name in bean_defs:
        check_cycle(full_name)

    return all_errors


def _resolve_root_dir():
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(os.path.abspath(sys.executable))
        if os.path.basename(exe_dir) == 'data':
            return os.path.dirname(exe_dir)
        return exe_dir
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

ROOT_DIR = _resolve_root_dir()
DATA_DIR = os.path.join(ROOT_DIR, 'data')
JSON_OUTPUT_DIR = os.path.join(ROOT_DIR, 'demo', 'data')
CS_OUTPUT_DIR = os.path.join(ROOT_DIR, 'demo', 'scripts', 'data', 'generated')

BASIC_TYPES = {'int', 'float', 'double', 'string', 'bool'}
LIST_PATTERN = re.compile(r'^\(list#sep=(.)\),(.+)$')
BEAN_FULL_NAME_RE = re.compile(r'^(\w+)\.(\w+)$')


def parse_type(type_str, bean_defs=None):
    type_str = type_str.strip()
    match = LIST_PATTERN.match(type_str)
    if match:
        elem_type = match.group(2).strip()
        return {
            'kind': 'list',
            'separator': match.group(1),
            'element_type': elem_type
        }
    if type_str in BASIC_TYPES:
        return {
            'kind': 'basic',
            'type': type_str
        }
    if bean_defs and type_str in bean_defs:
        return {
            'kind': 'bean',
            'type': type_str
        }
    if BEAN_FULL_NAME_RE.match(type_str):
        return {
            'kind': 'bean',
            'type': type_str
        }
    return {
        'kind': 'enum',
        'type': type_str
    }


def convert_basic_value(raw_value, basic_type):
    if raw_value is None:
        return None
    try:
        if basic_type == 'int':
            return int(float(str(raw_value)))
        elif basic_type == 'float':
            return float(str(raw_value))
        elif basic_type == 'double':
            return float(str(raw_value))
        elif basic_type == 'string':
            return str(raw_value)
        elif basic_type == 'bool':
            return str(raw_value).lower() in ('true', '1', 'yes')
    except (ValueError, TypeError) as e:
        raise ValueError(f"无法将 '{raw_value}' 转换为 {basic_type}: {e}")
    return raw_value


def convert_value(raw_value, type_info):
    if raw_value is None:
        return None
    if type_info['kind'] == 'list':
        raw_str = str(raw_value)
        if not raw_str:
            return []
        parts = raw_str.split(type_info['separator'])
        elem_type = type_info['element_type']
        if elem_type in BASIC_TYPES:
            return [convert_basic_value(p.strip(), elem_type) for p in parts]
        return [p.strip() for p in parts]
    if type_info['kind'] == 'basic':
        return convert_basic_value(raw_value, type_info['type'])
    if type_info['kind'] == 'enum':
        return str(raw_value).strip()
    return raw_value


def bean_col_span(bean_full_name, bean_defs):
    bean_def = bean_defs.get(bean_full_name)
    if bean_def is None:
        short_map = {fn.split('.')[-1]: fn for fn in bean_defs}
        bean_def = bean_defs.get(short_map.get(bean_full_name, ''))
    if bean_def is None:
        return 1
    total = 0
    for field in bean_def['fields']:
        field_type_info = parse_type(field['type'], bean_defs)
        if field_type_info['kind'] == 'bean':
            total += bean_col_span(field_type_info['type'], bean_defs)
        elif field_type_info['kind'] == 'list' and field_type_info['element_type'] not in BASIC_TYPES:
            total += bean_col_span(field_type_info['element_type'], bean_defs)
        else:
            total += 1
    return total


def bean_nesting_depth(bean_full_name, bean_defs):
    bean_def = bean_defs.get(bean_full_name)
    if bean_def is None:
        short_map = {fn.split('.')[-1]: fn for fn in bean_defs}
        bean_def = bean_defs.get(short_map.get(bean_full_name, ''))
    if bean_def is None:
        return 0
    max_depth = 0
    for field in bean_def['fields']:
        field_type_info = parse_type(field['type'], bean_defs)
        if field_type_info['kind'] == 'bean':
            depth = 1 + bean_nesting_depth(field_type_info['type'], bean_defs)
            max_depth = max(max_depth, depth)
        elif field_type_info['kind'] == 'list' and field_type_info['element_type'] not in BASIC_TYPES:
            depth = 1 + bean_nesting_depth(field_type_info['element_type'], bean_defs)
            max_depth = max(max_depth, depth)
    return max_depth


def build_bean_col_layout(bean_full_name, bean_defs, start_col, type_annotation_rows, comment_row, current_depth=0):
    bean_def = bean_defs.get(bean_full_name)
    if bean_def is None:
        short_map = {fn.split('.')[-1]: fn for fn in bean_defs}
        bean_def = bean_defs.get(short_map.get(bean_full_name, ''))
    if bean_def is None:
        return []
    layout = []
    col_offset = start_col
    for field in bean_def['fields']:
        field_type_info = parse_type(field['type'], bean_defs)
        if field_type_info['kind'] == 'bean':
            sub_layout = build_bean_col_layout(field_type_info['type'], bean_defs, col_offset, type_annotation_rows, comment_row, current_depth + 1)
            layout.append({
                'kind': 'bean_field',
                'name': field['name'],
                'type_info': field_type_info,
                'sub_layout': sub_layout,
                'col_count': sum(sl.get('col_count', 1) for sl in sub_layout),
            })
            col_offset += sum(sl.get('col_count', 1) for sl in sub_layout)
        elif field_type_info['kind'] == 'list' and field_type_info['element_type'] not in BASIC_TYPES:
            sub_layout = build_bean_col_layout(field_type_info['element_type'], bean_defs, col_offset, type_annotation_rows, comment_row, current_depth + 1)
            layout.append({
                'kind': 'list_bean_field',
                'name': field['name'],
                'type_info': field_type_info,
                'sub_layout': sub_layout,
                'col_count': sum(sl.get('col_count', 1) for sl in sub_layout),
            })
            col_offset += sum(sl.get('col_count', 1) for sl in sub_layout)
        else:
            if current_depth < len(type_annotation_rows):
                type_row = type_annotation_rows[current_depth]
                sub_name = type_row[col_offset] if col_offset < len(type_row) else ''
            else:
                sub_name = ''
            sub_comment = comment_row[col_offset] if col_offset < len(comment_row) else ''
            if not sub_name:
                sub_name = field['name']
            layout.append({
                'kind': 'basic_field',
                'name': field['name'],
                'type_info': field_type_info,
                'col_idx': col_offset,
                'col_count': 1,
                'sub_name': sub_name,
                'comment': sub_comment,
            })
            col_offset += 1
    return layout


def read_bean_row(layout, ws, row, bean_defs, enum_values, errors, sheet_name):
    result = {}
    has_data = False
    for entry in layout:
        if entry['kind'] == 'basic_field':
            raw_val = ws.cell(row=row, column=entry['col_idx'] + 1).value
            if raw_val is not None:
                has_data = True
                field_type_info = entry['type_info']
                if field_type_info['kind'] == 'enum':
                    enum_name = field_type_info['type']
                    if enum_name not in enum_values:
                        enum_values[enum_name] = set()
                    enum_values[enum_name].add(str(raw_val).strip())
                try:
                    result[entry['name']] = convert_value(raw_val, field_type_info)
                except ValueError as e:
                    errors.append(f"[错误] 表 {sheet_name} 第{row}行 字段 {entry['name']}: {e}")
                    result[entry['name']] = None
        elif entry['kind'] == 'bean_field':
            sub_result, sub_has = read_bean_row(entry['sub_layout'], ws, row, bean_defs, enum_values, errors, sheet_name)
            if sub_has:
                has_data = True
                result[entry['name']] = sub_result
        elif entry['kind'] == 'list_bean_field':
            sub_result, sub_has = read_bean_row(entry['sub_layout'], ws, row, bean_defs, enum_values, errors, sheet_name)
            if sub_has:
                has_data = True
                if entry['name'] not in result:
                    result[entry['name']] = []
                result[entry['name']].append(sub_result)
    return result, has_data


def process_sheet(ws, sheet_name, bean_defs=None):
    if ws.max_row < 3 or ws.max_column < 2:
        return None

    if bean_defs is None:
        bean_defs = {}

    errors = []

    row1 = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        row1.append(str(val).strip() if val is not None else '')

    row2 = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        row2.append(str(val).strip() if val is not None else '')

    max_nesting = 0
    i = 1
    while i < len(row1):
        header = row1[i]
        if header and header.startswith('*'):
            bean_type_str = row2[i] if i < len(row2) else ''
            bean_type_info = parse_type(bean_type_str, bean_defs)
            if bean_type_info['kind'] == 'bean':
                bean_full_name = bean_type_info['type']
            elif bean_type_info['kind'] == 'list' and bean_type_info['element_type'] not in BASIC_TYPES:
                bean_full_name = bean_type_info['element_type']
            else:
                bean_full_name = bean_type_str
            depth = bean_nesting_depth(bean_full_name, bean_defs)
            max_nesting = max(max_nesting, depth)
        i += 1

    header_rows = 3 + (1 + max_nesting) if max_nesting > 0 else 3
    data_start_row = header_rows + 1
    comment_row_idx = header_rows

    type_annotation_rows = []
    for r in range(3, comment_row_idx):
        row_data = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=col).value
            row_data.append(str(val).strip() if val is not None else '')
        type_annotation_rows.append(row_data)

    comment_row = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=comment_row_idx, column=col).value
        comment_row.append(str(val).strip() if val is not None else '')

    col_groups = []
    i = 1
    while i < len(row1):
        header = row1[i]
        if not header:
            i += 1
            continue

        if header.startswith('*'):
            bean_field_name = header[1:]
            bean_type_str = row2[i] if i < len(row2) else ''
            bean_type_info = parse_type(bean_type_str, bean_defs)

            if bean_type_info['kind'] == 'bean':
                bean_full_name = bean_type_info['type']
            elif bean_type_info['kind'] == 'list' and bean_type_info['element_type'] not in BASIC_TYPES:
                bean_full_name = bean_type_info['element_type']
            else:
                bean_full_name = bean_type_str

            bean_def = bean_defs.get(bean_full_name)
            if bean_def is None:
                short_map = {fn.split('.')[-1]: fn for fn in bean_defs}
                bean_def = bean_defs.get(short_map.get(bean_full_name, ''))

            if bean_def is None:
                errors.append(f"[错误] 表 {sheet_name} 第{i+1}列: Bean '{bean_full_name}' 未在 __bean__.xlsx 中定义")
                i += 1
                continue

            col_count = bean_col_span(bean_full_name, bean_defs)
            layout = build_bean_col_layout(bean_full_name, bean_defs, i, type_annotation_rows, comment_row)

            col_groups.append({
                'kind': 'bean',
                'field_name': bean_field_name,
                'bean_full_name': bean_full_name,
                'bean_def': bean_def,
                'type_str': bean_type_str,
                'layout': layout,
                'start_col': i,
                'col_count': col_count,
            })
            i += col_count
        else:
            type_str = row2[i] if i < len(row2) else ''
            type_info = parse_type(type_str, bean_defs)
            if type_info['kind'] == 'list':
                elem = type_info['element_type']
                if elem not in BASIC_TYPES and not VALID_IDENTIFIER_RE.match(elem):
                    errors.append(f"[错误] 表 {sheet_name} 第2行第{i+1}列: list元素类型 '{elem}' 不是合法标识符")

            comment = comment_row[i] if i < len(comment_row) else ''
            col_groups.append({
                'kind': 'basic',
                'field_name': header,
                'type_str': type_str,
                'type_info': type_info,
                'col_idx': i,
                'comment': comment,
            })
            i += 1

    if not col_groups:
        return None

    data_rows = []
    enum_values = {}
    current_record = None

    basic_groups = [g for g in col_groups if g['kind'] == 'basic']
    bean_groups = [g for g in col_groups if g['kind'] == 'bean']

    for row in range(data_start_row, ws.max_row + 1):
        marker = ws.cell(row=row, column=1).value
        if marker is not None and str(marker).strip() == '##':
            has_other_data = False
            for col in range(2, ws.max_column + 1):
                if ws.cell(row=row, column=col).value is not None:
                    has_other_data = True
                    break
            if not has_other_data:
                continue

        has_basic_data = False
        for g in basic_groups:
            raw_val = ws.cell(row=row, column=g['col_idx'] + 1).value
            if raw_val is not None:
                has_basic_data = True
                break

        has_bean_data = False
        for g in bean_groups:
            bean_data, bean_has = read_bean_row(g['layout'], ws, row, bean_defs, enum_values, errors, sheet_name)
            if bean_has:
                has_bean_data = True
                break

        if not has_basic_data and not has_bean_data:
            continue

        if has_basic_data:
            if current_record is not None:
                data_rows.append(current_record)
            current_record = {}

            for group in col_groups:
                if group['kind'] == 'basic':
                    col_idx = group['col_idx']
                    raw_val = ws.cell(row=row, column=col_idx + 1).value
                    type_info = group['type_info']
                    if type_info['kind'] == 'enum':
                        if raw_val is not None:
                            enum_name = type_info['type']
                            if enum_name not in enum_values:
                                enum_values[enum_name] = set()
                            enum_values[enum_name].add(str(raw_val).strip())
                    elif type_info['kind'] == 'list' and type_info['element_type'] not in BASIC_TYPES:
                        if raw_val is not None:
                            enum_name = type_info['element_type']
                            raw_str = str(raw_val).strip()
                            if raw_str:
                                if enum_name not in enum_values:
                                    enum_values[enum_name] = set()
                                parts = raw_str.split(type_info['separator'])
                                for p in parts:
                                    v = p.strip()
                                    if v:
                                        enum_values[enum_name].add(v)
                    try:
                        current_record[group['field_name']] = convert_value(raw_val, type_info)
                    except ValueError as e:
                        col_letter = chr(ord('A') + col_idx)
                        errors.append(f"[错误] 表 {sheet_name} 第{row}行 {col_letter}列 ({group['field_name']}): {e}")
                        current_record[group['field_name']] = None

                elif group['kind'] == 'bean':
                    bean_type_info = parse_type(group['type_str'], bean_defs)
                    if bean_type_info['kind'] == 'list':
                        current_record[group['field_name']] = []
                    else:
                        current_record[group['field_name']] = None

        if current_record is not None:
            for group in bean_groups:
                bean_type_info = parse_type(group['type_str'], bean_defs)
                bean_data, bean_has = read_bean_row(group['layout'], ws, row, bean_defs, enum_values, errors, sheet_name)
                if not bean_has:
                    continue

                if bean_type_info['kind'] == 'list':
                    if group['field_name'] not in current_record:
                        current_record[group['field_name']] = []
                    current_record[group['field_name']].append(bean_data)
                else:
                    current_record[group['field_name']] = bean_data

    if current_record is not None:
        data_rows.append(current_record)

    fields = []
    for group in col_groups:
        if group['kind'] == 'basic':
            fields.append({
                'name': group['field_name'],
                'type': group['type_str'] or 'string',
                'comment': group.get('comment', ''),
            })
        elif group['kind'] == 'bean':
            fields.append({
                'name': group['field_name'],
                'type': group['type_str'],
                'comment': '',
            })

    return {
        'tableName': sheet_name,
        'fields': fields,
        'data': data_rows,
        'enum_values': enum_values,
        'errors': errors
    }


def write_json(table_data, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    output = {
        'tableName': table_data['tableName'],
        'fields': table_data['fields'],
        'data': table_data['data']
    }
    json_path = os.path.join(output_dir, f"{table_data['tableName']}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  [JSON] {json_path}")


def generate_enum_cs(enum_defs, auto_enum_values):
    lines = []
    lines.append('// 自动生成的代码 - 请勿手动修改')
    lines.append('// 枚举定义：由导表工具根据 __enum__.xlsx 和数据表自动生成')
    lines.append('')

    all_enum_names = set(enum_defs.keys()) | set(auto_enum_values.keys())

    for enum_name in sorted(all_enum_names):
        if enum_name in enum_defs:
            members = enum_defs[enum_name]
            lines.append(f'public enum {enum_name}')
            lines.append('{')
            for i, m in enumerate(members):
                comma = ',' if i < len(members) - 1 else ''
                comment = f'  // {m["comment"]}' if m['comment'] else ''
                lines.append(f'    {m["name"]} = {m["value"]}{comma}{comment}')
            lines.append('}')
            lines.append('')
        else:
            values = sorted(auto_enum_values[enum_name])
            lines.append(f'public enum {enum_name}')
            lines.append('{')
            for i, val in enumerate(values):
                comma = ',' if i < len(values) - 1 else ''
                lines.append(f'    {val}{comma}')
            lines.append('}')
            lines.append('')

    return '\n'.join(lines)


def generate_bean_converter_cs(bean_defs):
    all_bean_short_names = {fn.split('.')[-1]: fn for fn in bean_defs}

    lines = []
    lines.append('// 自动生成的代码 - 请勿手动修改')
    lines.append('// Bean转换器：将Godot字典转换为Bean对象')
    lines.append('')
    lines.append('using Godot;')
    lines.append('using System;')
    lines.append('using System.Collections.Generic;')
    lines.append('')
    lines.append('public static class BeanConverter')
    lines.append('{')

    for full_name in sorted(bean_defs.keys()):
        bean_def = bean_defs[full_name]
        short_name = full_name.split('.')[-1]
        lines.append(f'    public static {short_name} To{short_name}(Godot.Collections.Dictionary dict)')
        lines.append(f'    {{')
        lines.append(f'        var bean = new {short_name}();')
        for f in bean_def['fields']:
            ftype = f['type']
            ftype_info = parse_type(ftype, bean_defs)
            if ftype_info['kind'] == 'basic':
                cs_read = {
                    'int': 'AsInt32()',
                    'float': 'AsSingle()',
                    'double': 'AsDouble()',
                    'string': 'AsString()',
                    'bool': 'AsBool()',
                }.get(ftype, 'AsString()')
                lines.append(f'        if (dict.ContainsKey("{f["name"]}")) bean.{f["name"]} = dict["{f["name"]}"].{cs_read};')
            elif ftype_info['kind'] == 'enum':
                lines.append(f'        if (dict.ContainsKey("{f["name"]}")) bean.{f["name"]} = Enum.Parse<{ftype}>(dict["{f["name"]}"].AsString());')
            elif ftype_info['kind'] == 'list':
                elem = ftype_info['element_type']
                elem_info = parse_type(elem, bean_defs)
                if elem in {'int', 'float', 'double', 'string', 'bool'}:
                    cs_conv = {
                        'int': 'AsInt32()',
                        'float': 'AsSingle()',
                        'double': 'AsDouble()',
                        'string': 'AsString()',
                        'bool': 'AsBool()',
                    }.get(elem, 'AsString()')
                    lines.append(f'        if (dict.ContainsKey("{f["name"]}"))')
                    lines.append(f'        {{')
                    lines.append(f'            var arr = dict["{f["name"]}"].AsGodotArray();')
                    lines.append(f'            var list = new List<{elem}>();')
                    lines.append(f'            foreach (var item in arr) list.Add(item.{cs_conv});')
                    lines.append(f'            bean.{f["name"]} = list;')
                    lines.append(f'        }}')
                elif elem_info['kind'] == 'bean':
                    elem_short = elem.split('.')[-1] if '.' in elem else elem
                    lines.append(f'        if (dict.ContainsKey("{f["name"]}"))')
                    lines.append(f'        {{')
                    lines.append(f'            var arr = dict["{f["name"]}"].AsGodotArray();')
                    lines.append(f'            var list = new List<{elem_short}>();')
                    lines.append(f'            foreach (var item in arr)')
                    lines.append(f'                list.Add(To{elem_short}(item.AsGodotDictionary()));')
                    lines.append(f'            bean.{f["name"]} = list;')
                    lines.append(f'        }}')
                else:
                    lines.append(f'        if (dict.ContainsKey("{f["name"]}"))')
                    lines.append(f'        {{')
                    lines.append(f'            var arr = dict["{f["name"]}"].AsGodotArray();')
                    lines.append(f'            var list = new List<string>();')
                    lines.append(f'            foreach (var item in arr) list.Add(item.AsString());')
                    lines.append(f'            bean.{f["name"]} = list;')
                    lines.append(f'        }}')
            elif ftype_info['kind'] == 'bean':
                ref_short = ftype.split('.')[-1]
                lines.append(f'        if (dict.ContainsKey("{f["name"]}")) bean.{f["name"]} = To{ref_short}(dict["{f["name"]}"].AsGodotDictionary());')
        lines.append(f'        return bean;')
        lines.append(f'    }}')
        lines.append('')

    lines.append('    public static T FromDict<T>(Godot.Collections.Dictionary dict) where T : new()')
    lines.append('    {')
    lines.append('        var typeName = typeof(T).Name;')
    lines.append('        return typeName switch')
    lines.append('        {')
    for full_name in sorted(bean_defs.keys()):
        short_name = full_name.split('.')[-1]
        lines.append(f'            "{short_name}" => (T)(object)To{short_name}(dict),')
    lines.append('            _ => new T()')
    lines.append('        };')
    lines.append('    }')
    lines.append('}')
    lines.append('')

    return '\n'.join(lines)


def generate_bean_cs(bean_defs):
    all_bean_short_names = {fn.split('.')[-1]: fn for fn in bean_defs}

    def resolve_cs_type(type_str):
        type_info = parse_type(type_str, bean_defs)
        if type_info['kind'] == 'basic':
            return type_info['type']
        elif type_info['kind'] == 'enum':
            return type_info['type']
        elif type_info['kind'] == 'bean':
            ref = type_info['type']
            if ref in bean_defs:
                return ref.split('.')[-1]
            if ref in all_bean_short_names:
                return ref
            return ref
        elif type_info['kind'] == 'list':
            elem = type_info['element_type']
            elem_cs = resolve_cs_type(elem)
            return f'List<{elem_cs}>'
        return type_str

    lines = []
    lines.append('// 自动生成的代码 - 请勿手动修改')
    lines.append('// 结构体定义：由导表工具根据 __bean__.xlsx 自动生成')
    lines.append('')
    lines.append('using Godot;')
    lines.append('using System;')
    lines.append('using System.Collections.Generic;')
    lines.append('')

    for full_name in sorted(bean_defs.keys()):
        bean_def = bean_defs[full_name]
        short_name = full_name.split('.')[-1]
        lines.append(f'public class {short_name}')
        lines.append('{')
        for f in bean_def['fields']:
            cs_type = resolve_cs_type(f['type'])
            lines.append(f'    public {cs_type} {f["name"]};')
        lines.append('')
        lines.append(f'    public override string ToString()')
        lines.append(f'    {{')
        parts = [f'{{{f["name"]}}}' for f in bean_def['fields']]
        lines.append(f'        return $"{short_name}({", ".join(parts)})";')
        lines.append(f'    }}')
        lines.append('}')
        lines.append('')

    return '\n'.join(lines)


def generate_table_record_cs():
    return '''// 自动生成的代码 - 请勿手动修改

using Godot;
using System;
using System.Collections.Generic;

public class TableRecord
{
    private readonly Dictionary<string, object> _fields = new();

    public string TableName { get; }

    public TableRecord(string tableName)
    {
        TableName = tableName;
    }

    internal void SetField(string name, object value)
    {
        _fields[name] = value;
    }

    public bool HasField(string name) => _fields.ContainsKey(name);

    public object GetRaw(string name)
    {
        if (_fields.TryGetValue(name, out var value))
            return value;
        GD.PrintErr($"[TableRecord] 字段不存在: {TableName}.{name}");
        return null;
    }

    public int GetInt(string name) => Convert.ToInt32(GetRaw(name));
    public float GetFloat(string name) => Convert.ToSingle(GetRaw(name));
    public double GetDouble(string name) => Convert.ToDouble(GetRaw(name));
    public string GetString(string name) => Convert.ToString(GetRaw(name)) ?? "";
    public bool GetBool(string name) => Convert.ToBoolean(GetRaw(name));

    public T GetEnum<T>(string name) where T : struct, Enum
    {
        var raw = GetRaw(name);
        if (raw is T typed)
            return typed;
        return Enum.Parse<T>(raw?.ToString() ?? "");
    }

    public List<int> GetIntList(string name) => GetList<int>(name);
    public List<float> GetFloatList(string name) => GetList<float>(name);
    public List<string> GetStringList(string name) => GetList<string>(name);

    public List<T> GetList<T>(string name)
    {
        var raw = GetRaw(name);
        if (raw is List<T> list)
            return list;
        return new List<T>();
    }

    public T GetBean<T>(string name) where T : new()
    {
        var raw = GetRaw(name);
        if (raw is T typed)
            return typed;
        if (raw is Godot.Collections.Dictionary dict)
            return BeanConverter.FromDict<T>(dict);
        return new T();
    }

    public List<T> GetBeanList<T>(string name) where T : new()
    {
        var raw = GetRaw(name);
        if (raw is List<T> list)
            return list;
        if (raw is Godot.Collections.Array arr)
        {
            var result = new List<T>();
            foreach (var item in arr)
            {
                if (item is Godot.Collections.Dictionary dict)
                    result.Add(BeanConverter.FromDict<T>(dict));
            }
            return result;
        }
        return new List<T>();
    }

    public override string ToString()
    {
        var parts = new List<string>();
        foreach (var kv in _fields)
            parts.Add($"{kv.Key}={kv.Value}");
        return $"[{TableName}] {{ {string.Join(", ", parts)} }}";
    }
}
'''


def generate_table_data_cs():
    return '''// 自动生成的代码 - 请勿手动修改

using Godot;
using System;
using System.Collections.Generic;

public class TableData
{
    private readonly List<TableRecord> _records = new();
    private readonly Dictionary<string, Dictionary<object, List<TableRecord>>> _indexes = new();
    private readonly Dictionary<string, string> _fieldTypes = new();

    public string TableName { get; }
    public int Count => _records.Count;
    public bool IsLoaded { get; private set; }

    public TableData(string tableName)
    {
        TableName = tableName;
    }

    public void Load()
    {
        var jsonText = FileAccess.GetFileAsString($"res://data/{TableName}.json");
        if (string.IsNullOrEmpty(jsonText))
        {
            GD.PrintErr($"[TableData] 加载表失败: {TableName}");
            IsLoaded = false;
            return;
        }
        var parsed = Json.ParseString(jsonText);
        if (parsed == null || parsed.VariantType != Variant.Type.Dictionary)
        {
            GD.PrintErr($"[TableData] JSON格式无效: {TableName}");
            IsLoaded = false;
            return;
        }
        var root = new Godot.Collections.Dictionary(parsed);

        _fieldTypes.Clear();
        var fieldsArray = root["fields"].AsGodotArray();
        foreach (var entry in fieldsArray)
        {
            var fieldDict = new Godot.Collections.Dictionary(entry);
            var fieldName = fieldDict["name"].AsString();
            var fieldType = fieldDict["type"].AsString();
            _fieldTypes[fieldName] = fieldType;
        }

        var dataArray = root["data"].AsGodotArray();
        foreach (var entry in dataArray)
        {
            var dict = new Godot.Collections.Dictionary(entry);
            var record = new TableRecord(TableName);

            foreach (var keyObj in dict.Keys)
            {
                var fieldName = keyObj.ToString();
                var variant = dict[keyObj];
                var fieldType = _fieldTypes.GetValueOrDefault(fieldName, "string");
                object value = ConvertVariant(variant, fieldType);
                record.SetField(fieldName, value);
            }

            _records.Add(record);
        }

        BuildIndexes();
        IsLoaded = true;
        GD.Print($"[TableData] 加载表 {TableName}: {_records.Count} 条记录");
    }

    private object ConvertVariant(Variant variant, string fieldType)
    {
        var typeInfo = ParseFieldType(fieldType);

        if (typeInfo.Kind == "list")
        {
            if (typeInfo.IsBeanList)
            {
                var arr = variant.AsGodotArray();
                var result = new List<Godot.Collections.Dictionary>();
                foreach (var item in arr)
                    result.Add(item.AsGodotDictionary());
                return result;
            }
            if (typeInfo.IsEnumList)
            {
                var arr = variant.AsGodotArray();
                var list = new List<string>();
                foreach (var item in arr)
                    list.Add(item.AsString());
                return list;
            }
            var arr2 = variant.AsGodotArray();
            var list2 = CreateTypedList(typeInfo.ElementType);
            foreach (var item in arr2)
            {
                list2.Add(ConvertSingleVariant(item, typeInfo.ElementType));
            }
            return list2;
        }

        if (typeInfo.Kind == "bean")
        {
            return variant.AsGodotDictionary();
        }

        return ConvertSingleVariant(variant, typeInfo.Kind);
    }

    private object ConvertSingleVariant(Variant variant, string type)
    {
        return type switch
        {
            "int" => variant.AsInt32(),
            "float" => variant.AsSingle(),
            "double" => variant.AsDouble(),
            "bool" => variant.AsBool(),
            "string" => variant.AsString(),
            _ => variant.AsString()
        };
    }

    private dynamic CreateTypedList(string elementType)
    {
        return elementType switch
        {
            "int" => new List<int>(),
            "float" => new List<float>(),
            "double" => new List<double>(),
            "bool" => new List<bool>(),
            "string" => new List<string>(),
            _ => new List<string>()
        };
    }

    private (string Kind, string ElementType, bool IsBeanList, bool IsEnumList) ParseFieldType(string fieldType)
    {
        var match = System.Text.RegularExpressions.Regex.Match(
            fieldType, @"^\\(list#sep=.\\),(.+)$");
        if (match.Success)
        {
            var elem = match.Groups[1].Value;
            var isBean = elem.Contains(".");
            var isEnum = !isBean && !new[] { "int", "float", "double", "string", "bool" }.Contains(elem);
            return ("list", elem, isBean, isEnum);
        }

        if (fieldType == "int" || fieldType == "float" || fieldType == "double"
            || fieldType == "string" || fieldType == "bool")
            return (fieldType, "", false, false);

        if (fieldType.Contains("."))
            return ("bean", fieldType, false, false);

        return (fieldType, "", false, false);
    }

    private void BuildIndexes()
    {
        _indexes.Clear();
        if (_records.Count == 0) return;

        foreach (var fieldName in _fieldTypes.Keys)
        {
            var typeInfo = ParseFieldType(_fieldTypes[fieldName]);
            if (typeInfo.Kind == "list" || typeInfo.Kind == "bean") continue;

            var index = new Dictionary<object, List<TableRecord>>();
            foreach (var record in _records)
            {
                var key = record.GetRaw(fieldName);
                if (key == null) continue;

                var boxKey = key is Enum ? key.ToString() : key;
                if (!index.ContainsKey(boxKey))
                    index[boxKey] = new List<TableRecord>();
                index[boxKey].Add(record);
            }
            _indexes[fieldName] = index;
        }
    }

    public TableRecord Find(string fieldName, object value)
    {
        var boxValue = value is Enum ? value.ToString() : value;
        if (_indexes.TryGetValue(fieldName, out var index))
        {
            if (index.TryGetValue(boxValue, out var list) && list.Count > 0)
                return list[0];
        }
        return null;
    }

    public List<TableRecord> FindAll(string fieldName, object value)
    {
        var boxValue = value is Enum ? value.ToString() : value;
        if (_indexes.TryGetValue(fieldName, out var index))
        {
            if (index.TryGetValue(boxValue, out var list))
                return new List<TableRecord>(list);
        }
        return new List<TableRecord>();
    }

    public TableRecord GetAt(int index)
    {
        if (index >= 0 && index < _records.Count)
            return _records[index];
        return null;
    }

    public List<TableRecord> GetAll() => new(_records);

    public string GetFieldType(string fieldName)
    {
        return _fieldTypes.GetValueOrDefault(fieldName, "");
    }
}
'''


def generate_table_manager_cs():
    return '''// 自动生成的代码 - 请勿手动修改

using Godot;
using System.Collections.Generic;

public partial class TableManager : Node
{
    public static TableManager Instance { get; private set; }

    private readonly Dictionary<string, TableData> _tables = new();
    private readonly List<string> _loadErrors = new();

    public override void _Ready()
    {
        Instance = this;
        LoadAllTables();

        if (_loadErrors.Count > 0)
        {
            GD.PrintErr($"[TableManager] {_loadErrors.Count} 张表加载失败!");
            foreach (var err in _loadErrors)
                GD.PrintErr($"  {err}");
            ShowLoadErrorDialog();
        }
        else
        {
            GD.Print($"[TableManager] 所有表加载完成, 共 {_tables.Count} 张表");
        }
    }

    private void LoadAllTables()
    {
        var dir = DirAccess.Open("res://data/");
        if (dir == null)
        {
            GD.PrintErr("[TableManager] 无法打开 res://data/ 目录");
            OS.Alert("无法打开数据目录 res://data/，请检查游戏安装是否完整", "数据加载错误");
            return;
        }

        dir.ListDirBegin();
        var fileName = dir.GetNext();
        while (fileName != string.Empty)
        {
            if (fileName.EndsWith(".json"))
            {
                var tableName = fileName.Substring(0, fileName.Length - 5);
                var table = new TableData(tableName);
                table.Load();
                _tables[tableName] = table;
                if (!table.IsLoaded)
                    _loadErrors.Add($"表 [{tableName}] 加载失败，请检查 res://data/{fileName} 是否存在且格式正确");
            }
            fileName = dir.GetNext();
        }
        dir.ListDirEnd();
    }

    private void ShowLoadErrorDialog()
    {
        var errorList = string.Join("\\n", _loadErrors);
        OS.Alert($"以下数据表加载失败，请重新导表后重启游戏:\\n\\n{errorList}", "数据加载错误");
    }

    public TableData GetTable(string tableName)
    {
        if (_tables.TryGetValue(tableName, out var table))
            return table;
        GD.PrintErr($"[TableManager] 表不存在: {tableName}");
        return null;
    }

}
'''


def scan_xlsx_files(data_dir):
    xlsx_files = []
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~'):
                xlsx_files.append(os.path.join(root, f))
    return xlsx_files


def to_pascal_case(name):
    if '_' in name:
        return ''.join(part.capitalize() for part in name.split('_') if part)
    return name[0].upper() + name[1:] if name else name


def resolve_wrapper_cs_type(type_str, bean_defs):
    type_info = parse_type(type_str, bean_defs)
    if type_info['kind'] == 'basic':
        return type_info['type']
    elif type_info['kind'] == 'enum':
        return type_info['type']
    elif type_info['kind'] == 'bean':
        ref = type_info['type']
        all_bean_short = {fn.split('.')[-1]: fn for fn in bean_defs}
        if ref in bean_defs:
            return ref.split('.')[-1]
        if ref in all_bean_short:
            return ref
        return ref
    elif type_info['kind'] == 'list':
        elem = type_info['element_type']
        elem_cs = resolve_wrapper_cs_type(elem, bean_defs)
        return f'List<{elem_cs}>'
    return type_str


def get_field_accessor(type_str, field_name, bean_defs):
    type_info = parse_type(type_str, bean_defs)
    if type_info['kind'] == 'basic':
        accessor_map = {
            'int': 'GetInt',
            'float': 'GetFloat',
            'double': 'GetDouble',
            'string': 'GetString',
            'bool': 'GetBool',
        }
        method = accessor_map.get(type_info['type'], 'GetString')
        return f'_raw.{method}("{field_name}")'
    elif type_info['kind'] == 'enum':
        return f'_raw.GetEnum<{type_info["type"]}>("{field_name}")'
    elif type_info['kind'] == 'bean':
        ref = type_info['type']
        short = ref.split('.')[-1] if '.' in ref else ref
        return f'_raw.GetBean<{short}>("{field_name}")'
    elif type_info['kind'] == 'list':
        elem = type_info['element_type']
        elem_info = parse_type(elem, bean_defs)
        if elem in {'int', 'float', 'double', 'string', 'bool'}:
            method_map = {
                'int': 'GetIntList',
                'float': 'GetFloatList',
                'string': 'GetStringList',
            }
            method = method_map.get(elem)
            if method:
                return f'_raw.{method}("{field_name}")'
            return f'_raw.GetList<{elem}>("{field_name}")'
        elif elem_info['kind'] == 'enum':
            return f'_raw.GetStringList("{field_name}").ConvertAll(s => Enum.Parse<{elem}>(s))'
        elif elem_info['kind'] == 'bean':
            short = elem.split('.')[-1] if '.' in elem else elem
            return f'_raw.GetBeanList<{short}>("{field_name}")'
        else:
            return f'_raw.GetList<string>("{field_name}")'
    return f'_raw.GetString("{field_name}")'


def generate_tables_cs(all_tables, bean_defs):
    lines = []
    lines.append('// 自动生成的代码 - 请勿手动修改')
    lines.append('// 强类型表包装：由导表工具根据数据表自动生成')
    lines.append('')
    lines.append('using Godot;')
    lines.append('using System;')
    lines.append('using System.Collections.Generic;')
    lines.append('')

    for table in all_tables:
        table_name = table['tableName']
        class_name = to_pascal_case(table_name)
        fields = table['fields']

        lines.append(f'public class {class_name}Row')
        lines.append('{')
        lines.append(f'    private readonly TableRecord _raw;')
        lines.append(f'    public {class_name}Row(TableRecord raw) {{ _raw = raw; }}')
        lines.append('')

        for field in fields:
            prop_name = to_pascal_case(field['name'])
            cs_type = resolve_wrapper_cs_type(field['type'], bean_defs)
            accessor = get_field_accessor(field['type'], field['name'], bean_defs)
            lines.append(f'    public {cs_type} {prop_name} => {accessor};')

        lines.append('}')
        lines.append('')

    for table in all_tables:
        table_name = table['tableName']
        class_name = to_pascal_case(table_name)
        row_class = f'{class_name}Row'
        fields = table['fields']

        lines.append(f'public class {class_name}Table')
        lines.append('{')
        lines.append(f'    private readonly TableData _raw;')
        lines.append(f'    private List<{row_class}> _rows;')
        lines.append('')
        lines.append(f'    public {class_name}Table(TableData raw) {{ _raw = raw; }}')
        lines.append(f'    public int Count => _raw.Count;')
        lines.append('')

        lines.append(f'    public List<{row_class}> GetAll()')
        lines.append(f'    {{')
        lines.append(f'        if (_rows == null)')
        lines.append(f'            _rows = _raw.GetAll().ConvertAll(r => new {row_class}(r));')
        lines.append(f'        return _rows;')
        lines.append(f'    }}')
        lines.append('')

        has_id = len(fields) > 0 and fields[0]['name'] == 'id' and fields[0]['type'] == 'int'
        if has_id:
            lines.append(f'    public {row_class} FindById(int id)')
            lines.append(f'    {{')
            lines.append(f'        var record = _raw.Find("id", id);')
            lines.append(f'        return record != null ? new {row_class}(record) : null;')
            lines.append(f'    }}')
            lines.append('')

        lines.append(f'    public {row_class} Find(string fieldName, object value)')
        lines.append(f'    {{')
        lines.append(f'        var record = _raw.Find(fieldName, value);')
        lines.append(f'        return record != null ? new {row_class}(record) : null;')
        lines.append(f'    }}')
        lines.append('')

        lines.append(f'    public List<{row_class}> FindAll(string fieldName, object value)')
        lines.append(f'    {{')
        lines.append(f'        return _raw.FindAll(fieldName, value).ConvertAll(r => new {row_class}(r));')
        lines.append(f'    }}')

        lines.append('}')
        lines.append('')

    lines.append('public static class Tables')
    lines.append('{')

    for table in all_tables:
        table_name = table['tableName']
        class_name = to_pascal_case(table_name)
        table_class = f'{class_name}Table'
        prop_name = class_name

        lines.append(f'    private static {table_class} _{table_name};')
        lines.append(f'    public static {table_class} {prop_name}')
        lines.append(f'    {{')
        lines.append(f'        get')
        lines.append(f'        {{')
        lines.append(f'            if (_{table_name} == null)')
        lines.append(f'                _{table_name} = new {table_class}(TableManager.Instance.GetTable("{table_name}"));')
        lines.append(f'            return _{table_name};')
        lines.append(f'        }}')
        lines.append(f'    }}')
        lines.append('')

    lines.append('}')
    lines.append('')

    return '\n'.join(lines)


def main():
    print("=" * 60)
    print("《通货与铁砧》导表工具 v1.0")
    print("=" * 60)

    if not os.path.exists(DATA_DIR):
        print(_c(f"[错误] 数据目录不存在: {DATA_DIR}", _RED))
        sys.exit(1)

    xlsx_files = scan_xlsx_files(DATA_DIR)
    if not xlsx_files:
        print(_c("[警告] 未找到任何 xlsx 文件", _YELLOW))
        sys.exit(0)

    enum_files = []
    bean_files = []
    data_files = []
    for f in xlsx_files:
        name = os.path.splitext(os.path.basename(f))[0]
        if name == '__bean__':
            bean_files.append(f)
        elif name.startswith('__') and name.endswith('__'):
            enum_files.append(f)
        else:
            data_files.append(f)

    print(f"\n扫描到 {len(data_files)} 个数据表, {len(enum_files)} 个枚举表, {len(bean_files)} 个结构体表:")
    for f in data_files:
        print(f"  - {os.path.relpath(f, ROOT_DIR)}")
    for f in enum_files:
        print(f"  - {os.path.relpath(f, ROOT_DIR)} (枚举表)")
    for f in bean_files:
        print(f"  - {os.path.relpath(f, ROOT_DIR)} (结构体表)")

    all_errors = []
    enum_defs = {}

    for xlsx_path in enum_files:
        file_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        print(f"\n解析枚举表: {file_name}.xlsx")
        defs, errors = parse_enum_xlsx(xlsx_path)
        all_errors.extend(errors)
        for enum_name, members in defs.items():
            if enum_name in enum_defs:
                all_errors.append(f"[错误] 枚举 '{enum_name}' 在多个枚举表中重复定义")
            else:
                enum_defs[enum_name] = members
                member_names = [m['name'] for m in members]
                print(f"  枚举 {enum_name}: {', '.join(member_names)}")

    bean_defs = {}

    for xlsx_path in bean_files:
        file_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        print(f"\n解析结构体表: {file_name}.xlsx")
        defs, errors = parse_bean_xlsx(xlsx_path)
        all_errors.extend(errors)
        for full_name, bean_def in defs.items():
            if full_name in bean_defs:
                all_errors.append(f"[错误] Bean '{full_name}' 在多个结构体表中重复定义")
            else:
                bean_defs[full_name] = bean_def
                field_names = [f['name'] for f in bean_def['fields']]
                print(f"  结构体 {full_name}: {', '.join(field_names)}")

    if bean_defs:
        ref_errors = validate_bean_refs(bean_defs, enum_defs)
        all_errors.extend(ref_errors)

    all_tables = []
    all_enum_values = {}

    for xlsx_path in data_files:
        file_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        print(f"\n处理: {file_name}.xlsx")

        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            print(_c(f"  [错误] 无法打开文件: {e}", _RED))
            continue

        data_sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 4 or ws.max_column < 2:
                continue
            first_cell = ws.cell(row=1, column=1).value
            first_cell_str = str(first_cell).strip() if first_cell else ''
            if first_cell_str.endswith('字段名称'):
                data_sheets.append((sheet_name, ws))

        if len(data_sheets) == 1:
            table_name = file_name
            sheet_name, ws = data_sheets[0]
            result = process_sheet(ws, table_name, bean_defs)
            if result:
                all_tables.append(result)
                all_errors.extend(result.get('errors', []))
                for enum_name, values in result['enum_values'].items():
                    if enum_name not in all_enum_values:
                        all_enum_values[enum_name] = set()
                    all_enum_values[enum_name].update(values)
                print(f"  Sheet '{sheet_name}' → 表 '{table_name}' ({len(result['data'])} 条记录)")
        elif len(data_sheets) > 1:
            for sheet_name, ws in data_sheets:
                table_name = f"{file_name}_{sheet_name}"
                result = process_sheet(ws, table_name, bean_defs)
                if result:
                    all_tables.append(result)
                    all_errors.extend(result.get('errors', []))
                    for enum_name, values in result['enum_values'].items():
                        if enum_name not in all_enum_values:
                            all_enum_values[enum_name] = set()
                        all_enum_values[enum_name].update(values)
                    print(f"  Sheet '{sheet_name}' → 表 '{table_name}' ({len(result['data'])} 条记录)")

        wb.close()

    for enum_name in enum_defs:
        if enum_name in all_enum_values:
            del all_enum_values[enum_name]

    for enum_name in all_enum_values:
        if enum_name in enum_defs:
            defined_members = {m['name'] for m in enum_defs[enum_name]}
            for val in all_enum_values[enum_name]:
                if val not in defined_members:
                    all_errors.append(f"[错误] 数据表中使用了枚举 {enum_name} 的成员 '{val}'，但该成员未在 __enum__.xlsx 中定义 (已有: {', '.join(sorted(defined_members))})")

    if all_errors:
        print(f"\n{'=' * 60}")
        print(_c(f"发现 {len(all_errors)} 个错误:", _RED))
        for err in all_errors:
            print(_c(f"  {err}", _RED))
        print(_c(f"\n请修复以上错误后重新导出!", _RED))
        print(f"{'=' * 60}")
        sys.exit(1)

    if not all_tables and not enum_defs:
        print(_c("\n[警告] 未找到有效的数据表或枚举定义", _YELLOW))
        sys.exit(0)

    for xlsx_path in enum_files + bean_files:
        stale_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        stale_json = os.path.join(JSON_OUTPUT_DIR, f"{stale_name}.json")
        if os.path.exists(stale_json):
            os.remove(stale_json)
            print(f"  [清理] 删除残留文件: {os.path.relpath(stale_json, ROOT_DIR)}")

    print(f"\n{'=' * 60}")
    print("导出 JSON 文件:")
    for table in all_tables:
        write_json(table, JSON_OUTPUT_DIR)

    print(f"\n{'=' * 60}")
    print("生成 C# 代码:")

    os.makedirs(CS_OUTPUT_DIR, exist_ok=True)

    if enum_defs or all_enum_values:
        enum_code = generate_enum_cs(enum_defs, all_enum_values)
        enum_path = os.path.join(CS_OUTPUT_DIR, '__enum__.cs')
        with open(enum_path, 'w', encoding='utf-8') as f:
            f.write(enum_code)
        print(f"  [C#] {enum_path}")
        for enum_name, members in sorted(enum_defs.items()):
            member_names = [m['name'] for m in members]
            print(f"        枚举 {enum_name}: {', '.join(member_names)} (显式定义)")
        for enum_name, values in sorted(all_enum_values.items()):
            print(f"        枚举 {enum_name}: {', '.join(sorted(values))} (自动推断)")

    if bean_defs:
        bean_code = generate_bean_cs(bean_defs)
        bean_path = os.path.join(CS_OUTPUT_DIR, '__bean__.cs')
        with open(bean_path, 'w', encoding='utf-8') as f:
            f.write(bean_code)
        print(f"  [C#] {bean_path}")
        for full_name in sorted(bean_defs.keys()):
            short_name = full_name.split('.')[-1]
            field_names = [f['name'] for f in bean_defs[full_name]['fields']]
            print(f"        结构体 {short_name}: {', '.join(field_names)}")

        converter_code = generate_bean_converter_cs(bean_defs)
        converter_path = os.path.join(CS_OUTPUT_DIR, 'BeanConverter.cs')
        with open(converter_path, 'w', encoding='utf-8') as f:
            f.write(converter_code)
        print(f"  [C#] {converter_path}")

    record_code = generate_table_record_cs()
    record_path = os.path.join(CS_OUTPUT_DIR, 'TableRecord.cs')
    with open(record_path, 'w', encoding='utf-8') as f:
        f.write(record_code)
    print(f"  [C#] {record_path}")

    table_code = generate_table_data_cs()
    table_path = os.path.join(CS_OUTPUT_DIR, 'TableData.cs')
    with open(table_path, 'w', encoding='utf-8') as f:
        f.write(table_code)
    print(f"  [C#] {table_path}")

    manager_code = generate_table_manager_cs()
    manager_path = os.path.join(CS_OUTPUT_DIR, 'TableManager.cs')
    with open(manager_path, 'w', encoding='utf-8') as f:
        f.write(manager_code)
    print(f"  [C#] {manager_path}")

    if all_tables:
        tables_code = generate_tables_cs(all_tables, bean_defs)
        tables_path = os.path.join(CS_OUTPUT_DIR, 'Tables.cs')
        with open(tables_path, 'w', encoding='utf-8') as f:
            f.write(tables_code)
        print(f"  [C#] {tables_path}")
        for table in all_tables:
            class_name = to_pascal_case(table['tableName'])
            field_names = [f['name'] for f in table['fields']]
            print(f"        {class_name}Row / {class_name}Table: {', '.join(field_names)}")

    print(f"\n{'=' * 60}")
    print(_c(f"导出完成! 共处理 {len(all_tables)} 张表, {len(enum_defs)} 个枚举, {len(bean_defs)} 个结构体", _GREEN))
    print(f"JSON 输出: {os.path.relpath(JSON_OUTPUT_DIR, ROOT_DIR)}/")
    print(f"C# 输出:   {os.path.relpath(CS_OUTPUT_DIR, ROOT_DIR)}/")
    fixed_files = ['__enum__.cs', '__bean__.cs', 'BeanConverter.cs', 'TableRecord.cs', 'TableData.cs', 'TableManager.cs', 'Tables.cs']
    print(f"  固定文件: {' + '.join(fixed_files)}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    try:
        main()
    except SystemExit:
        pass
    finally:
        if getattr(sys, 'frozen', False):
            import os
            in_ide = any(key.startswith('VSCODE_') or key.startswith('IDEA_') or key == 'TERM_PROGRAM' for key in os.environ)
            in_terminal = os.environ.get('WT_SESSION') or os.environ.get('TERM')
            if not in_ide and not in_terminal:
                input("\n按任意键关闭...")
